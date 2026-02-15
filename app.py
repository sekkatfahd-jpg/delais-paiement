import streamlit as st
import pandas as pd
from datetime import datetime
import io
import json
import os
import shutil
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Délais de Paiement", layout="wide")

# Dossier de cache pour les fichiers uploadés
CACHE_DIR = os.path.join(os.path.dirname(__file__), "cache_files")
CACHE_GL_FILE = os.path.join(CACHE_DIR, "grand_livre.xlsx")
CACHE_BALANCE_FILE = os.path.join(CACHE_DIR, "balance.xlsx")

# Créer le dossier de cache s'il n'existe pas
if not os.path.exists(CACHE_DIR):
    os.makedirs(CACHE_DIR)

def save_uploaded_file(uploaded_file, cache_path):
    """Sauvegarder un fichier uploadé dans le cache"""
    try:
        with open(cache_path, 'wb') as f:
            f.write(uploaded_file.getbuffer())
        return True
    except Exception:
        return False

def load_cached_file(cache_path):
    """Charger un fichier depuis le cache"""
    try:
        if os.path.exists(cache_path):
            with open(cache_path, 'rb') as f:
                return f.read()
    except Exception:
        pass
    return None

# CSS personnalisé pour une interface plus professionnelle
st.markdown("""
<style>
    /* Style général */
    .main {
        background-color: #f8f9fa;
    }

    /* Titre principal */
    .main-title {
        background: linear-gradient(90deg, #1e3a5f 0%, #2d5a87 100%);
        color: white;
        padding: 1.5rem 2rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }

    .main-title h1 {
        margin: 0;
        font-size: 1.8rem;
        font-weight: 600;
    }

    .main-title p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        font-size: 1rem;
    }

    /* Cartes de configuration */
    .config-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        border: 1px solid #e9ecef;
        margin-bottom: 1rem;
    }

    .config-card h3 {
        color: #1e3a5f;
        margin-bottom: 1rem;
        font-size: 1.1rem;
        border-bottom: 2px solid #2d5a87;
        padding-bottom: 0.5rem;
    }

    /* Métriques */
    .metric-container {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        text-align: center;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        border-left: 4px solid #2d5a87;
    }

    /* Boutons */
    .stButton > button {
        background: linear-gradient(90deg, #1e3a5f 0%, #2d5a87 100%);
        color: white;
        border: none;
        padding: 0.75rem 2rem;
        font-weight: 600;
        border-radius: 8px;
        transition: all 0.3s ease;
    }

    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(30, 58, 95, 0.3);
    }

    /* File uploader */
    .stFileUploader {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        border: 2px dashed #dee2e6;
    }

    /* Expander */
    .streamlit-expanderHeader {
        background: white;
        border-radius: 8px;
    }

    /* Dataframe */
    .stDataFrame {
        border-radius: 8px;
        overflow: hidden;
    }

    /* Section divider */
    .section-divider {
        height: 3px;
        background: linear-gradient(90deg, #1e3a5f 0%, #2d5a87 50%, #1e3a5f 100%);
        border-radius: 2px;
        margin: 2rem 0;
    }

    /* Footer */
    .footer {
        background: #1e3a5f;
        color: white;
        padding: 1rem;
        border-radius: 8px;
        text-align: center;
        margin-top: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# Titre principal stylisé
st.markdown("""
<div class="main-title">
    <h1>Declaration relative aux delais de paiement</h1>
    <p>Rapprochement Factures - Paiements</p>
</div>
""", unsafe_allow_html=True)

# Fichier de configuration pour persister les journaux
CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config_journaux.json")

def load_config():
    """Charger la configuration des journaux depuis le fichier"""
    default_config = {
        "journaux_achat": "ACHAT\nACH",
        "journaux_banque": "BANQUE\nBNQ\nCHEQUE"
    }
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return default_config

def save_config(journaux_achat, journaux_banque):
    """Sauvegarder la configuration des journaux"""
    try:
        config = {
            "journaux_achat": journaux_achat,
            "journaux_banque": journaux_banque
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# Charger la configuration sauvegardée
config = load_config()

# Interface centrée avec colonnes
st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

col_left, col_center, col_right = st.columns([1, 2, 1])

with col_center:
    st.markdown('<div class="config-card">', unsafe_allow_html=True)
    st.markdown("### Fichiers requis")

    # Initialiser session_state pour les fichiers cachés
    if 'gl_loaded_from_cache' not in st.session_state:
        st.session_state.gl_loaded_from_cache = False
    if 'balance_loaded_from_cache' not in st.session_state:
        st.session_state.balance_loaded_from_cache = False

    col_file1, col_file2 = st.columns(2)
    with col_file1:
        grand_livre_file = st.file_uploader(
            "Grand Livre (Excel)",
            type=['xlsx', 'xls'],
            help="Colonnes : A=Date, B=Journal, C=Compte, D=N°Pièce, E=Libellé, F=Mvt, G=Facture, I=Lettrage"
        )
        # Sauvegarder le fichier uploadé dans le cache
        if grand_livre_file is not None:
            save_uploaded_file(grand_livre_file, CACHE_GL_FILE)
            grand_livre_file.seek(0)  # Remettre le curseur au début
        # Charger depuis le cache si aucun fichier uploadé
        elif os.path.exists(CACHE_GL_FILE):
            cached_gl = load_cached_file(CACHE_GL_FILE)
            if cached_gl:
                grand_livre_file = io.BytesIO(cached_gl)
                grand_livre_file.name = "grand_livre.xlsx"
                if not st.session_state.gl_loaded_from_cache:
                    st.info("Grand Livre chargé depuis le cache")
                    st.session_state.gl_loaded_from_cache = True

    with col_file2:
        balance_file = st.file_uploader(
            "Balance des Comptes (Excel)",
            type=['xlsx', 'xls'],
            help="Pour récupérer les noms des fournisseurs"
        )
        # Sauvegarder le fichier uploadé dans le cache
        if balance_file is not None:
            save_uploaded_file(balance_file, CACHE_BALANCE_FILE)
            balance_file.seek(0)  # Remettre le curseur au début
        # Charger depuis le cache si aucun fichier uploadé
        elif os.path.exists(CACHE_BALANCE_FILE):
            cached_balance = load_cached_file(CACHE_BALANCE_FILE)
            if cached_balance:
                balance_file = io.BytesIO(cached_balance)
                balance_file.name = "balance.xlsx"
                if not st.session_state.balance_loaded_from_cache:
                    st.info("Balance chargée depuis le cache")
                    st.session_state.balance_loaded_from_cache = True

    # Bouton pour effacer le cache
    if os.path.exists(CACHE_GL_FILE) or os.path.exists(CACHE_BALANCE_FILE):
        if st.button("🗑️ Effacer les fichiers en cache", type="secondary"):
            if os.path.exists(CACHE_GL_FILE):
                os.remove(CACHE_GL_FILE)
            if os.path.exists(CACHE_BALANCE_FILE):
                os.remove(CACHE_BALANCE_FILE)
            st.session_state.gl_loaded_from_cache = False
            st.session_state.balance_loaded_from_cache = False
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="config-card">', unsafe_allow_html=True)
    st.markdown("### Configuration des Journaux")

    col_j1, col_j2 = st.columns(2)
    with col_j1:
        journaux_achat_input = st.text_area(
            "Journaux d'Achat (un par ligne)",
            value=config.get("journaux_achat", "ACHAT\nACH"),
            height=100,
            key="journaux_achat_input"
        )
        journaux_achat = [j.strip() for j in journaux_achat_input.split('\n') if j.strip()]

    with col_j2:
        journaux_banque_input = st.text_area(
            "Journaux de Banque (un par ligne)",
            value=config.get("journaux_banque", "BANQUE\nBNQ\nCHEQUE"),
            height=100,
            key="journaux_banque_input"
        )
        journaux_banque = [j.strip() for j in journaux_banque_input.split('\n') if j.strip()]

    st.markdown('</div>', unsafe_allow_html=True)

# Sauvegarder la configuration si elle a changé
if journaux_achat_input != config.get("journaux_achat") or journaux_banque_input != config.get("journaux_banque"):
    save_config(journaux_achat_input, journaux_banque_input)

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Fonction pour charger le grand livre (sans en-tête)
@st.cache_data
def load_grand_livre(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    colonnes = {
        0: 'Date',           # A
        1: 'Journal',        # B
        2: 'Compte',         # C
        3: 'NumPiece',       # D
        4: 'Libelle',        # E
        5: 'MontantMvt',     # F
        6: 'MontantFacture', # G
        8: 'Lettrage'        # I
    }

    cols_to_keep = {k: v for k, v in colonnes.items() if k < len(df.columns)}
    df = df[list(cols_to_keep.keys())].copy()
    df.columns = list(cols_to_keep.values())

    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

    # Créer les colonnes manquantes avec des zéros si elles n'existent pas
    if 'MontantMvt' not in df.columns:
        df['MontantMvt'] = 0
    else:
        df['MontantMvt'] = pd.to_numeric(df['MontantMvt'], errors='coerce').fillna(0)

    if 'MontantFacture' not in df.columns:
        df['MontantFacture'] = 0
    else:
        df['MontantFacture'] = pd.to_numeric(df['MontantFacture'], errors='coerce').fillna(0)

    # Vérifier si la colonne Lettrage existe
    if 'Lettrage' not in df.columns:
        df['Lettrage'] = ''
    else:
        df['Lettrage'] = df['Lettrage'].astype(str).str.strip()
        df.loc[df['Lettrage'] == 'nan', 'Lettrage'] = ''

    df['Compte'] = df['Compte'].astype(str).str.strip()
    df['Compte'] = df['Compte'].apply(lambda x: x[:-2] if x.endswith('.0') else x)

    df['NumPiece'] = df['NumPiece'].astype(str).str.strip()
    df['NumPiece'] = df['NumPiece'].apply(lambda x: x[:-2] if x.endswith('.0') else x)
    df.loc[df['NumPiece'] == 'nan', 'NumPiece'] = ''

    # Nettoyer la colonne Journal (enlever espaces)
    if 'Journal' in df.columns:
        df['Journal'] = df['Journal'].astype(str).str.strip()
        df.loc[df['Journal'] == 'nan', 'Journal'] = ''

    return df

@st.cache_data
def load_balance(file_bytes):
    df_with_header = pd.read_excel(io.BytesIO(file_bytes), header=0)
    first_col = df_with_header.columns[0]
    if isinstance(first_col, (int, float)) or (isinstance(first_col, str) and first_col.replace('.', '').isdigit()):
        df = pd.read_excel(io.BytesIO(file_bytes), header=None)
        return df, False
    return df_with_header, True

def creer_dict_fournisseurs(balance_df, has_header):
    dict_fournisseurs = {}
    col_compte = None
    col_nom = None

    if has_header:
        cols = balance_df.columns.tolist()
        for col in cols:
            col_str = str(col).lower()
            if any(x in col_str for x in ['compte', 'n°', 'numero', 'code', 'num']):
                col_compte = col
                break
        for col in cols:
            col_str = str(col).lower()
            if any(x in col_str for x in ['nom', 'intitule', 'intitulé', 'libelle', 'libellé', 'raison', 'designation', 'désignation']):
                col_nom = col
                break

    if col_compte is None or col_nom is None:
        if len(balance_df.columns) >= 2:
            col_compte = balance_df.columns[0]
            col_nom = balance_df.columns[1]

    if col_compte is not None and col_nom is not None:
        for _, row in balance_df.iterrows():
            compte_raw = row[col_compte]
            if pd.isna(compte_raw):
                continue
            compte = str(compte_raw).strip()
            if compte.endswith('.0'):
                compte = compte[:-2]
            nom = str(row[col_nom]).strip() if pd.notna(row[col_nom]) else ''
            if compte and compte != 'nan' and nom and nom != 'nan':
                dict_fournisseurs[compte] = nom

    return dict_fournisseurs, col_compte, col_nom

def generer_nouvelle_lettre(lettres_utilisees):
    """Génère une nouvelle lettre de lettrage non utilisée"""
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # Essayer lettres simples
    for lettre in alphabet:
        if lettre not in lettres_utilisees:
            return lettre
    # Essayer combinaisons AA, AB, etc.
    for l1 in alphabet:
        for l2 in alphabet:
            combo = l1 + l2
            if combo not in lettres_utilisees:
                return combo
    # Essayer AAA, AAB, etc.
    for l1 in alphabet:
        for l2 in alphabet:
            for l3 in alphabet:
                combo = l1 + l2 + l3
                if combo not in lettres_utilisees:
                    return combo
    return "ZZZ"

def corriger_erreurs_lettrage(grand_livre_df, journaux_achat, journaux_banque):
    """
    Fonction désactivée - le lettrage du grand livre est utilisé tel quel.
    La correction automatique créait des lettrages artificiels qui cassaient
    l'affectation correcte des avoirs aux factures.
    """
    return {}

def traiter_rapprochement(grand_livre_df, dict_fournisseurs, journaux_achat, journaux_banque):

    journaux_connus = set(journaux_achat + journaux_banque)

    # Corriger les erreurs de lettrage
    corrections_lettrage = corriger_erreurs_lettrage(grand_livre_df, journaux_achat, journaux_banque)

    # Ajouter une colonne de lettrage corrigé au grand livre
    def get_lettrage_corrige(row):
        key = (row['Compte'], row['NumPiece'], row['Date'], row['MontantFacture'])
        if key in corrections_lettrage:
            return corrections_lettrage[key]
        key2 = (row['Compte'], row['NumPiece'], row['Date'], row['MontantMvt'])
        if key2 in corrections_lettrage:
            return corrections_lettrage[key2]
        return row['Lettrage']

    grand_livre_df = grand_livre_df.copy()
    grand_livre_df['LettrageCorrige'] = grand_livre_df.apply(get_lettrage_corrige, axis=1)

    # === DÉTECTION DES EFFETS À PAYER (comptes 4415) ===
    # Les effets à payer permettent de trouver la vraie date de paiement des factures
    # Flux: Facture 4411 -> Effet créé (4411 soldé, 4415 crédité) -> Paiement réel (4415 débité)
    
    dict_paiements_effet = {}  # Clé: (compte_4411, lettrage_4411) -> {date_paiement, montant}
    
    # Identifier les comptes 4415 (effets à payer)
    comptes_4415 = grand_livre_df[
        grand_livre_df['Compte'].astype(str).str.startswith('4415')
    ]['Compte'].unique()
    
    if len(comptes_4415) > 0:
        # Pour chaque effet créé sur 4415 (MontantFacture > 0), trouver le paiement correspondant
        effets_4415 = grand_livre_df[
            (grand_livre_df['Compte'].astype(str).str.startswith('4415')) &
            (grand_livre_df['MontantFacture'] > 0)
        ]
        
        for _, effet in effets_4415.iterrows():
            montant_effet = effet['MontantFacture']
            date_creation = effet['Date']
            lettrage_4415 = str(effet['LettrageCorrige']).strip()
            compte_4415 = str(effet['Compte']).strip()
            
            if not lettrage_4415:
                continue
                
            # Trouver le paiement réel sur 4415 (même compte, même lettrage, MontantMvt > 0)
            paiement_4415 = grand_livre_df[
                (grand_livre_df['Compte'] == compte_4415) &
                (grand_livre_df['LettrageCorrige'].astype(str).str.strip() == lettrage_4415) &
                (grand_livre_df['MontantMvt'] > 0)
            ]
            
            if len(paiement_4415) > 0:
                date_paiement_reel = paiement_4415.iloc[0]['Date']
                
                # Trouver le mouvement correspondant sur 4411 (même montant, même date que création effet)
                # C'est le mouvement qui solde le 4411 en créditant le 4415
                mvt_4411_effet = grand_livre_df[
                    (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
                    (grand_livre_df['MontantMvt'] == montant_effet) &
                    (grand_livre_df['Date'] == date_creation)
                ]
                
                if len(mvt_4411_effet) > 0:
                    compte_4411 = str(mvt_4411_effet.iloc[0]['Compte']).strip()
                    lettrage_4411 = str(mvt_4411_effet.iloc[0]['LettrageCorrige']).strip()
                    
                    if lettrage_4411:
                        # Stocker: pour ce groupe (compte, lettrage) sur 4411, 
                        # le paiement réel est à cette date
                        key = (compte_4411, lettrage_4411)
                        if key not in dict_paiements_effet:
                            dict_paiements_effet[key] = {
                                'date_paiement': date_paiement_reel,
                                'montant': montant_effet,
                                'date_effet': date_creation
                            }
                        else:
                            # Si plusieurs effets pour le même groupe, on cumule
                            dict_paiements_effet[key]['montant'] += montant_effet

    # Identifier les factures
    lignes_fournisseurs = grand_livre_df[
        (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
        (grand_livre_df['Journal'].isin(journaux_achat))
    ].copy()

    factures = lignes_fournisseurs[lignes_fournisseurs['MontantFacture'] != 0].copy()

    # Identifier les avoirs (avec détails pour affectation chronologique)
    avoirs = grand_livre_df[
        (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
        (grand_livre_df['Journal'].isin(journaux_achat)) &
        (grand_livre_df['MontantMvt'] > 0)
    ].copy()

    # Séparer avoirs lettrés et non lettrés
    avoirs_non_lettres = []
    
    # Dictionnaire des avoirs par (compte, lettrage) - liste avec dates pour tri chronologique
    avoirs_par_groupe = {}
    for _, avoir in avoirs.iterrows():
        compte = str(avoir['Compte']).strip()
        lettrage = str(avoir['LettrageCorrige']).strip()
        if lettrage:
            key = (compte, lettrage)
            if key not in avoirs_par_groupe:
                avoirs_par_groupe[key] = []
            avoirs_par_groupe[key].append({
                'date': avoir['Date'],
                'montant': avoir['MontantMvt'],
                'type': 'avoir',
                'avoir_obj': avoir  # Stocker l'objet complet pour accès aux détails
            })
        else:
            # Avoir sans lettrage - à traiter séparément
            avoirs_non_lettres.append(avoir)

    # Identifier les paiements
    paiements = grand_livre_df[
        (grand_livre_df['Journal'].isin(journaux_banque)) &
        (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
        (grand_livre_df['MontantMvt'] != 0)
    ].copy()

    # Identifier les remboursements fournisseurs (journal banque avec MontantFacture > 0)
    # Ces remboursements représentent un retour d'argent au fournisseur (avoir encaissé)
    remboursements_fournisseurs = grand_livre_df[
        (grand_livre_df['Journal'].isin(journaux_banque)) &
        (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
        (grand_livre_df['MontantFacture'] > 0)
    ].copy()

    # Identifier les remboursements fournisseurs non lettrés
    # Critères : Journal banque + Compte 4411 + MontantFacture > 0 + Lettrage vide
    remboursements_non_lettres = grand_livre_df[
        (grand_livre_df['Journal'].isin(journaux_banque)) &
        (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
        (grand_livre_df['MontantFacture'] > 0) &
        ((grand_livre_df['LettrageCorrige'].str.strip() == '') | (grand_livre_df['LettrageCorrige'].isna()))
    ].copy()

    # Dictionnaire des paiements par (compte, lettrage) - liste avec dates pour tri chronologique
    paiements_par_groupe = {}
    for _, paiement in paiements.iterrows():
        compte = str(paiement['Compte']).strip()
        lettrage = str(paiement['LettrageCorrige']).strip()
        if lettrage:
            key = (compte, lettrage)
            if key not in paiements_par_groupe:
                paiements_par_groupe[key] = []
            paiements_par_groupe[key].append({
                'date': paiement['Date'],
                'montant': abs(paiement['MontantMvt']),
                'type': 'paiement',
                'paiement_obj': paiement
            })

    # Ajouter les paiements par effet à paiements_par_groupe
    # Ces paiements ont une date de paiement réelle différente de la date de l'effet
    for key, effet_info in dict_paiements_effet.items():
        compte, lettrage = key
        if key not in paiements_par_groupe:
            paiements_par_groupe[key] = []
        paiements_par_groupe[key].append({
            'date': effet_info['date_paiement'],  # Date du paiement réel sur 4415
            'montant': effet_info['montant'],
            'type': 'paiement_effet',
            'paiement_obj': None  # Pas d'objet paiement, c'est un effet
        })

    # Ajouter les remboursements lettrés à paiements_par_groupe (avec montant négatif)
    # Un remboursement est un "paiement négatif" - le fournisseur nous rend de l'argent
    remboursements_par_groupe = {}
    for _, remb in remboursements_fournisseurs.iterrows():
        compte = str(remb['Compte']).strip()
        lettrage = str(remb['LettrageCorrige']).strip()
        if lettrage:
            key = (compte, lettrage)
            if key not in remboursements_par_groupe:
                remboursements_par_groupe[key] = []
            remboursements_par_groupe[key].append({
                'date': remb['Date'],
                'montant': remb['MontantFacture'],  # Montant du remboursement
                'type': 'remboursement',
                'remboursement_obj': remb
            })

    # Identifier les OD lettrées - Journal différent de achat/banque + Lettrage présent
    # On distingue deux types selon le CONTEXTE du groupe :
    # 1. Écarts de change : MontantMvt > 0 → toujours réparti sur les factures
    # 2. OD avec MontantFacture > 0 et MontantMvt = 0 :
    #    - Si le groupe a des PAIEMENTS → c'est une perte/gain de change → répartir
    #    - Si le groupe N'A PAS de paiements → c'est un reclassement → ligne séparée
    
    od_lignes_all = grand_livre_df[
        (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
        (~grand_livre_df['Journal'].isin(journaux_connus)) &
        ((grand_livre_df['MontantFacture'] > 0) | (grand_livre_df['MontantMvt'] > 0)) &
        (grand_livre_df['LettrageCorrige'].str.strip() != '') &
        (grand_livre_df['LettrageCorrige'].notna())
    ].copy()

    # Séparer les OD selon leur type
    # Type 1 : MontantMvt > 0 (écarts de change classiques)
    od_ecarts_change_mvt = od_lignes_all[od_lignes_all['MontantMvt'] > 0].copy()
    
    # Type 2 : MontantFacture > 0 et MontantMvt = 0 (à classifier selon contexte)
    od_montant_facture = od_lignes_all[
        (od_lignes_all['MontantFacture'] > 0) & 
        (od_lignes_all['MontantMvt'] == 0)
    ].copy()

    # Stocker les écarts de change (MontantMvt > 0) pour répartition
    # SAUF les mouvements vers 4415 (effets) qui sont traités comme des paiements
    dict_od_brut = {}
    for _, od in od_ecarts_change_mvt.iterrows():
        compte = str(od['Compte']).strip()
        lettrage = str(od['LettrageCorrige']).strip()
        if lettrage:
            key = (compte, lettrage)
            montant_od = abs(od['MontantMvt'])
            
            # Vérifier si c'est un effet (paiement via 4415)
            if key in dict_paiements_effet:
                # C'est un effet, pas une OD - déjà traité comme paiement
                continue
            
            if key not in dict_od_brut:
                dict_od_brut[key] = 0
            dict_od_brut[key] += montant_od
    
    # Grouper les factures par (compte, lettrage) - nécessaire AVANT la classification des OD
    factures_par_groupe = {}
    for _, facture in factures.iterrows():
        compte = str(facture['Compte']).strip()
        lettrage = str(facture['LettrageCorrige']).strip()
        key = (compte, lettrage)
        if key not in factures_par_groupe:
            factures_par_groupe[key] = []
        factures_par_groupe[key].append(facture)

    # Trier les factures par date dans chaque groupe
    for key in factures_par_groupe:
        factures_par_groupe[key] = sorted(
            factures_par_groupe[key],
            key=lambda x: x['Date'] if pd.notna(x['Date']) else pd.Timestamp.max
        )
    
    # Pour les OD avec MontantFacture > 0 : classifier selon le contexte du groupe
    # - Si groupe a des FACTURES ET des paiements → répartir sur factures
    # - Sinon (pas de factures ou pas de paiements) → reclassement
    od_reclassements_par_groupe = {}
    for _, od in od_montant_facture.iterrows():
        compte = str(od['Compte']).strip()
        lettrage = str(od['LettrageCorrige']).strip()
        if lettrage:
            key = (compte, lettrage)
            # Vérifier si ce groupe a des paiements bancaires ET des factures
            groupe_a_paiements = key in paiements_par_groupe and len(paiements_par_groupe[key]) > 0
            groupe_a_factures = key in factures_par_groupe and len(factures_par_groupe[key]) > 0
            
            if groupe_a_paiements and groupe_a_factures:
                # Le groupe a des factures ET des paiements → l'OD est une perte/gain de change à répartir
                montant_od = abs(od['MontantFacture'])
                if key not in dict_od_brut:
                    dict_od_brut[key] = 0
                dict_od_brut[key] += montant_od
            else:
                # Le groupe n'a PAS de factures ou PAS de paiements → c'est un reclassement
                if key not in od_reclassements_par_groupe:
                    od_reclassements_par_groupe[key] = []
                od_reclassements_par_groupe[key].append(od)

    # Créer le tableau de résultats
    resultats = []

    # Structure pour stocker les factures avec solde restant (pour affectation des paiements non lettrés)
    factures_solde_restant = []  # Liste de dicts avec infos facture + solde restant

    # Suivre les groupes déjà traités par affectation chronologique
    groupes_traites = set()

    # Traiter les groupes avec lettrage (affectation chronologique)
    for key in factures_par_groupe:
        compte_fournisseur, lettrage_corrige = key

        if not lettrage_corrige:
            continue  # Les non-lettrés seront traités séparément

        factures_groupe = factures_par_groupe[key]
        avoirs_groupe = avoirs_par_groupe.get(key, [])
        paiements_groupe = paiements_par_groupe.get(key, [])
        od_brut = dict_od_brut.get(key, 0)

        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")

        # Calculer totaux pour déterminer gain/perte
        total_factures = sum(abs(fac['MontantFacture']) for fac in factures_groupe)
        total_avoirs = sum(avoir['montant'] for avoir in avoirs_groupe)
        total_paiements_montant = sum(paiement['montant'] for paiement in paiements_groupe)
        # Soustraire les remboursements du total des paiements
        remboursements_groupe = remboursements_par_groupe.get(key, [])
        total_remboursements = sum(remb['montant'] for remb in remboursements_groupe)
        total_paiements_net = total_paiements_montant - total_remboursements

        # CORRECTION: Si une seule facture avec plusieurs paiements et OD, forcer le traitement proratisé
        if len(factures_groupe) == 1 and len(paiements_groupe) > 1 and od_brut > 0:
            # Forcer le passage par la logique de plusieurs paiements
            facture = factures_groupe[0]
            lettrage_original = str(facture['Lettrage']).strip()
            lettrage_corrige_affiche = lettrage_corrige if lettrage_corrige != lettrage_original else ''
            montant_original = abs(facture['MontantFacture'])
            
            total_paiements_facture = sum(p['montant'] for p in paiements_groupe)
            
            # Déterminer si c'est une perte de change (paiements > factures)
            is_perte_change_local = total_paiements_facture > (montant_original + 0.001)
            
            for paiement in paiements_groupe:
                montant_paiement = paiement['montant']
                ratio = montant_paiement / total_paiements_facture if total_paiements_facture > 0 else 0
                
                # Proratiser la facture et l'OD
                montant_facture_prorata = montant_original * ratio
                od_prorata = od_brut * ratio if is_perte_change_local else -od_brut * ratio
                solde_paiement = montant_facture_prorata - montant_paiement + od_prorata
                
                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_facture_prorata,
                    'Avoir': 0,
                    'Montant facture net': montant_facture_prorata,
                    'Date de paiement': paiement['date'],
                    'Montant du paiement': montant_paiement,
                    'OD': od_prorata,
                    'Montant du paiement groupé': 0,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': solde_paiement
                })
            
            groupes_traites.add(key)
            continue

        # Déterminer si c'est une perte de change (paiements > factures) ou un gain
        # Perte de change : on a payé PLUS que la facture → OD positif
        # Gain / Passage en gain : on a payé MOINS que la facture → OD négatif
        # Utiliser une petite tolérance pour éviter les erreurs d'arrondi
        is_perte_change = total_paiements_net > (total_factures - total_avoirs + 0.001)

        if od_brut > 0:
            if is_perte_change:
                montant_od_signe = od_brut  # Positif pour perte de change
            else:
                montant_od_signe = -od_brut  # Négatif pour gain
        else:
            montant_od_signe = 0

        # Combiner avoirs, paiements et remboursements et trier chronologiquement
        operations = []
        for avoir in avoirs_groupe:
            operations.append({
                'date': avoir['date'],
                'montant': avoir['montant'],
                'type': 'avoir',
                'montant_restant': avoir['montant']
            })
        for paiement in paiements_groupe:
            operations.append({
                'date': paiement['date'],
                'montant': paiement['montant'],
                'type': 'paiement',
                'montant_restant': paiement['montant'],
                'paiement_obj': paiement.get('paiement_obj')
            })
        # Ajouter les remboursements comme paiements négatifs
        remboursements_groupe = remboursements_par_groupe.get(key, [])
        for remb in remboursements_groupe:
            operations.append({
                'date': remb['date'],
                'montant': -remb['montant'],  # Négatif car c'est un remboursement
                'type': 'remboursement',
                'montant_restant': -remb['montant'],
                'remboursement_obj': remb.get('remboursement_obj')
            })

        # Trier par date (plus ancien d'abord)
        operations = sorted(
            operations,
            key=lambda x: x['date'] if pd.notna(x['date']) else pd.Timestamp.max
        )

        # Calculer le total des paiements pour "Montant du paiement groupé"
        # Inclure les remboursements (qui ont des montants négatifs)
        total_paiements = sum(op['montant'] for op in operations if op['type'] in ('paiement', 'remboursement'))

        # Préparer les factures avec leur solde restant
        factures_avec_solde = []
        for fac in factures_groupe:
            factures_avec_solde.append({
                'facture': fac,
                'montant_original': abs(fac['MontantFacture']),
                'solde_restant': abs(fac['MontantFacture']),
                'avoir_affecte': 0,
                'paiements_affectes': []
            })

        # Affectation chronologique des opérations aux factures
        # Les remboursements sont traités séparément après
        remboursements_a_traiter = []
        for op in operations:
            if op['type'] == 'remboursement':
                # Les remboursements sont traités séparément
                remboursements_a_traiter.append(op)
                continue
                
            montant_op_restant = op['montant_restant']

            for fac_info in factures_avec_solde:
                if montant_op_restant <= 0:
                    break
                if fac_info['solde_restant'] <= 0:
                    continue

                montant_a_affecter = min(montant_op_restant, fac_info['solde_restant'])

                if op['type'] == 'avoir':
                    fac_info['avoir_affecte'] += montant_a_affecter
                else:
                    fac_info['paiements_affectes'].append({
                        'montant': montant_a_affecter,
                        'date': op['date'],
                        'total_paiement': op['montant']
                    })

                fac_info['solde_restant'] -= montant_a_affecter
                montant_op_restant -= montant_a_affecter
        
        # Traiter les remboursements : ils correspondent aux avoirs
        # Chaque remboursement annule un avoir du même groupe
        for remb in remboursements_a_traiter:
            montant_remb = abs(remb['montant'])  # Montant positif pour comparaison
            # Ajouter le remboursement comme paiement négatif sur la première facture avec avoir
            for fac_info in factures_avec_solde:
                if fac_info['avoir_affecte'] > 0:
                    fac_info['paiements_affectes'].append({
                        'montant': -montant_remb,  # Négatif car remboursement
                        'date': remb['date'],
                        'total_paiement': -montant_remb,
                        'type': 'remboursement'
                    })
                    break

        # Cas spécial : Plusieurs factures + 1 paiement unique + OD (pénalité/perte de change OU gain de change)
        # Le paiement couvre toutes les factures avec un écart (OD)
        # Factures 1 à N-1 : paiement = montant facture, OD = 0, Solde = 0
        # Dernière facture : paiement = reste du paiement, OD = total OD (+ ou -), Solde = 0
        if (len(factures_groupe) > 1 and len(paiements_groupe) == 1
            and total_avoirs == 0 and od_brut > 0):

            # Trier les factures par date
            factures_triees = sorted(factures_avec_solde,
                key=lambda x: x['facture']['Date'] if pd.notna(x['facture']['Date']) else pd.Timestamp.max)

            paiement_info = paiements_groupe[0]
            date_paiement = paiement_info['date']
            paiement_total = paiement_info['montant']
            paiement_restant = paiement_total

            for idx, fac_info in enumerate(factures_triees):
                facture = fac_info['facture']
                lettrage_original = str(facture['Lettrage']).strip()
                lettrage_corrige_affiche = lettrage_corrige if lettrage_corrige != lettrage_original else ''
                montant_facture = fac_info['montant_original']

                is_derniere_facture = (idx == len(factures_triees) - 1)

                if is_derniere_facture:
                    # Dernière facture : reçoit le reste du paiement et tout l'OD
                    montant_paiement_affiche = paiement_restant
                    # OD positif si perte de change, négatif si gain de change
                    od_affiche = od_brut if is_perte_change else -od_brut
                else:
                    # Factures intermédiaires : paiement = montant facture, OD = 0, Solde = 0
                    montant_paiement_affiche = montant_facture
                    od_affiche = 0
                    paiement_restant -= montant_facture

                solde = montant_facture - montant_paiement_affiche + od_affiche

                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_facture,
                    'Avoir': 0,
                    'Montant facture net': montant_facture,
                    'Date de paiement': date_paiement,
                    'Montant du paiement': montant_paiement_affiche,
                    'OD': od_affiche,
                    'Montant du paiement groupé': paiement_total,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': solde
                })

            groupes_traites.add(key)
            continue  # Passer au groupe suivant, ne pas exécuter la boucle normale

        # Cas spécial : Plusieurs factures + plusieurs paiements + OD (écart de régularisation)
        # Rapprochement 1-1 entre factures et paiements, l'OD sur la première facture avec écart
        if (len(factures_groupe) > 1 and len(paiements_groupe) > 1
            and total_avoirs == 0 and od_brut > 0):

            # Trier les factures par date
            factures_triees = sorted(factures_avec_solde,
                key=lambda x: x['facture']['Date'] if pd.notna(x['facture']['Date']) else pd.Timestamp.max)

            # Créer une copie des paiements disponibles pour le rapprochement
            paiements_disponibles = [{'montant': p['montant'], 'date': p['date'], 'utilise': False}
                                     for p in paiements_groupe]

            od_deja_affecte = False

            for idx, fac_info in enumerate(factures_triees):
                facture = fac_info['facture']
                lettrage_original = str(facture['Lettrage']).strip()
                lettrage_corrige_affiche = lettrage_corrige if lettrage_corrige != lettrage_original else ''
                montant_facture = fac_info['montant_original']

                # Chercher le paiement correspondant (exact ou le plus proche)
                meilleur_paiement = None
                meilleur_ecart = float('inf')
                meilleur_idx = -1

                for p_idx, paiement in enumerate(paiements_disponibles):
                    if paiement['utilise']:
                        continue
                    ecart = abs(paiement['montant'] - montant_facture)
                    if ecart < meilleur_ecart:
                        meilleur_ecart = ecart
                        meilleur_paiement = paiement
                        meilleur_idx = p_idx

                if meilleur_paiement:
                    paiements_disponibles[meilleur_idx]['utilise'] = True
                    montant_paiement = meilleur_paiement['montant']
                    date_paiement = meilleur_paiement['date']

                    # L'OD est affecté à la première facture qui a un écart
                    if not od_deja_affecte and abs(montant_facture - montant_paiement) > 0.001:
                        # OD positif si perte de change, négatif si gain de change
                        od_affiche = od_brut if is_perte_change else -od_brut
                        od_deja_affecte = True
                    else:
                        od_affiche = 0

                    solde = montant_facture - montant_paiement + od_affiche
                else:
                    # Pas de paiement trouvé
                    montant_paiement = 0
                    date_paiement = None
                    od_affiche = 0
                    solde = montant_facture

                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_facture,
                    'Avoir': 0,
                    'Montant facture net': montant_facture,
                    'Date de paiement': date_paiement,
                    'Montant du paiement': montant_paiement,
                    'OD': od_affiche,
                    'Montant du paiement groupé': total_paiements,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': solde
                })

            groupes_traites.add(key)
            continue  # Passer au groupe suivant

        # Cas spécial : Plusieurs factures + PAS de paiement + OD (annulation par OD/effet)
        # L'OD annule les factures - répartir l'OD au prorata sur chaque facture
        if (len(factures_groupe) > 1 and len(paiements_groupe) == 0
            and total_avoirs == 0 and od_brut > 0):

            # Trier les factures par date
            factures_triees = sorted(factures_avec_solde,
                key=lambda x: x['facture']['Date'] if pd.notna(x['facture']['Date']) else pd.Timestamp.max)

            for fac_info in factures_triees:
                facture = fac_info['facture']
                lettrage_original = str(facture['Lettrage']).strip()
                lettrage_corrige_affiche = lettrage_corrige if lettrage_corrige != lettrage_original else ''
                montant_facture = fac_info['montant_original']

                # Calculer l'OD au prorata : OD_i = OD_total × (Montant_facture_i / Total_factures)
                ratio = montant_facture / total_factures if total_factures > 0 else 0
                od_prorata = od_brut * ratio

                # L'OD est négative (annule la facture) donc solde = montant - od_prorata = 0
                solde = montant_facture - od_prorata

                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_facture,
                    'Avoir': 0,
                    'Montant facture net': montant_facture,
                    'Date de paiement': None,
                    'Montant du paiement': 0,
                    'OD': -od_prorata,  # Négatif car annulation
                    'Montant du paiement groupé': 0,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': solde
                })

            groupes_traites.add(key)
            continue  # Passer au groupe suivant

        # Générer les lignes de résultat pour chaque facture
        for fac_info in factures_avec_solde:
            facture = fac_info['facture']
            lettrage_original = str(facture['Lettrage']).strip()
            lettrage_corrige_affiche = lettrage_corrige if lettrage_corrige != lettrage_original else ''

            avoir_total = fac_info['avoir_affecte']
            paiements_affectes = fac_info['paiements_affectes']
            montant_original = fac_info['montant_original']

            # Si la facture a des avoirs et des paiements
            if avoir_total > 0 and len(paiements_affectes) > 0:
                # Ligne pour l'avoir
                montant_facture_avoir = avoir_total
                montant_facture_net_avoir = 0  # Avoir couvre cette portion
                solde_avoir = montant_facture_net_avoir - 0 + montant_od_signe

                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_facture_avoir,
                    'Avoir': avoir_total,
                    'Montant facture net': montant_facture_net_avoir,
                    'Date de paiement': None,
                    'Montant du paiement': 0,
                    'OD': montant_od_signe,
                    'Montant du paiement groupé': 0,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': solde_avoir
                })

                # Lignes pour chaque paiement
                for paiement_aff in paiements_affectes:
                    montant_paiement = paiement_aff['montant']
                    montant_facture_affiche = montant_paiement
                    montant_facture_net_affiche = montant_facture_affiche
                    solde_paiement = montant_facture_net_affiche - montant_paiement + montant_od_signe

                    resultats.append({
                        'Date de facture': facture['Date'],
                        'N° de facture': facture['NumPiece'],
                        'N° compte fournisseur': compte_fournisseur,
                        'Nom du fournisseur': nom_fournisseur,
                        'Libellé de l\'opération': facture['Libelle'],
                        'Montant de la facture': montant_facture_affiche,
                        'Avoir': 0,
                        'Montant facture net': montant_facture_net_affiche,
                        'Date de paiement': paiement_aff['date'],
                        'Montant du paiement': montant_paiement,
                        'OD': montant_od_signe,
                        'Montant du paiement groupé': total_paiements if len(factures_groupe) > 1 else 0,
                        'Lettrage': lettrage_original,
                        'Lettrage corrigé': lettrage_corrige_affiche,
                        'Solde': solde_paiement
                    })

            # Si la facture n'a que des avoirs
            elif avoir_total > 0 and len(paiements_affectes) == 0:
                montant_facture_net = montant_original - avoir_total
                solde = montant_facture_net + montant_od_signe

                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_original,
                    'Avoir': avoir_total,
                    'Montant facture net': montant_facture_net,
                    'Date de paiement': None,
                    'Montant du paiement': 0,
                    'OD': montant_od_signe,
                    'Montant du paiement groupé': 0,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': solde
                })

            # Si la facture n'a que des paiements (pas d'avoir)
            elif avoir_total == 0 and len(paiements_affectes) > 0:
                # Cas spécial : Écart de change avec plusieurs paiements (perte de change)
                # On applique le prorata : MF_i = Facture × (Paiement_i / Total Paiements)
                # OD_i = OD × (Paiement_i / Total Paiements)
                if is_perte_change and len(paiements_affectes) > 1 and od_brut > 0:
                    # Prorata pour écart de change
                    total_paiements_facture = sum(p['montant'] for p in paiements_affectes)
                    for paiement_aff in paiements_affectes:
                        montant_paiement = paiement_aff['montant']
                        ratio = montant_paiement / total_paiements_facture if total_paiements_facture > 0 else 0
                        # Montant facture prorata
                        montant_facture_prorata = montant_original * ratio
                        # OD prorata (positif car perte de change)
                        od_prorata = od_brut * ratio
                        # Solde = MF - Paiement + OD = 0
                        solde_paiement = montant_facture_prorata - montant_paiement + od_prorata

                        resultats.append({
                            'Date de facture': facture['Date'],
                            'N° de facture': facture['NumPiece'],
                            'N° compte fournisseur': compte_fournisseur,
                            'Nom du fournisseur': nom_fournisseur,
                            'Libellé de l\'opération': facture['Libelle'],
                            'Montant de la facture': montant_facture_prorata,
                            'Avoir': 0,
                            'Montant facture net': montant_facture_prorata,
                            'Date de paiement': paiement_aff['date'],
                            'Montant du paiement': montant_paiement,
                            'OD': od_prorata,
                            'Montant du paiement groupé': 0,  # Pas de paiement groupé (plusieurs paiements pour 1 facture)
                            'Lettrage': lettrage_original,
                            'Lettrage corrigé': lettrage_corrige_affiche,
                            'Solde': solde_paiement
                        })
                # Plusieurs paiements sans écart de change spécial
                elif len(paiements_affectes) > 1:
                    # CORRECTION: Proratiser la facture quand plusieurs paiements
                    total_paiements_facture = sum(p['montant'] for p in paiements_affectes)
                    
                    for paiement_aff in paiements_affectes:
                        montant_paiement = paiement_aff['montant']
                        
                        # Proratiser la facture selon le ratio du paiement
                        ratio = montant_paiement / total_paiements_facture if total_paiements_facture > 0 else 0
                        montant_facture_affiche = montant_original * ratio
                        montant_facture_net_affiche = montant_facture_affiche
                        
                        # Proratiser l'OD aussi
                        od_affiche = montant_od_signe * ratio if montant_od_signe != 0 else 0
                        solde_paiement = montant_facture_net_affiche - montant_paiement + od_affiche

                        resultats.append({
                            'Date de facture': facture['Date'],
                            'N° de facture': facture['NumPiece'],
                            'N° compte fournisseur': compte_fournisseur,
                            'Nom du fournisseur': nom_fournisseur,
                            'Libellé de l\'opération': facture['Libelle'],
                            'Montant de la facture': montant_facture_affiche,
                            'Avoir': 0,
                            'Montant facture net': montant_facture_net_affiche,
                            'Date de paiement': paiement_aff['date'],
                            'Montant du paiement': montant_paiement,
                            'OD': od_affiche,
                            'Montant du paiement groupé': 0,  # CORRECTION: 0 car plusieurs paiements pour 1 facture
                            'Lettrage': lettrage_original,
                            'Lettrage corrigé': lettrage_corrige_affiche,
                            'Solde': solde_paiement
                        })
                else:
                    # Un seul paiement (potentiellement partiel)
                    paiement_aff = paiements_affectes[0]
                    montant_paiement = paiement_aff['montant']

                    # S'il y a un OD (gain OU perte de change)
                    if od_brut > 0:
                        montant_paiement = paiement_aff.get('total_paiement', paiement_aff['montant'])
                        montant_facture_affiche = montant_original
                        solde = montant_original - montant_paiement + montant_od_signe
                    else:
                        # Paiement partiel ou exact sans OD
                        montant_facture_affiche = montant_paiement
                        solde = 0

                    resultats.append({
                        'Date de facture': facture['Date'],
                        'N° de facture': facture['NumPiece'],
                        'N° compte fournisseur': compte_fournisseur,
                        'Nom du fournisseur': nom_fournisseur,
                        'Libellé de l\'opération': facture['Libelle'],
                        'Montant de la facture': montant_facture_affiche,
                        'Avoir': 0,
                        'Montant facture net': montant_facture_affiche,
                        'Date de paiement': paiement_aff['date'],
                        'Montant du paiement': montant_paiement,
                        'OD': montant_od_signe if od_brut > 0 else 0,
                        'Montant du paiement groupé': total_paiements if len(factures_groupe) > 1 else 0,
                        'Lettrage': lettrage_original,
                        'Lettrage corrigé': lettrage_corrige_affiche,
                        'Solde': solde
                    })

            # Si la facture n'a ni avoir ni paiement
            else:
                solde = montant_original + montant_od_signe

                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_original,
                    'Avoir': 0,
                    'Montant facture net': montant_original,
                    'Date de paiement': None,
                    'Montant du paiement': 0,
                    'OD': montant_od_signe,
                    'Montant du paiement groupé': 0,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': solde
                })

            # Stocker les factures avec solde restant pour affectation des paiements non lettrés
            # IMPORTANT: Vérifier le solde RÉEL après OD pour ne pas affecter les paiements
            # à des factures déjà soldées par leur groupe lettré
            if fac_info['solde_restant'] > 0.01:
                # Calculer le solde réel en tenant compte de l'OD
                solde_reel = fac_info['solde_restant'] + montant_od_signe
                
                # N'ajouter QUE si le solde réel > 0 (la facture a vraiment besoin d'un paiement)
                if solde_reel > 0.01:
                    factures_solde_restant.append({
                        'facture': facture,
                        'compte': compte_fournisseur,
                        'nom_fournisseur': nom_fournisseur,
                        'solde_restant': solde_reel,  # Solde APRÈS OD
                        'lettrage_original': lettrage_original,
                        'lettrage_corrige_affiche': lettrage_corrige_affiche,
                        'od_signe': 0  # OD déjà comptabilisé dans solde_restant
                    })

        # Traiter les avoirs non affectés (quand la facture était déjà payée)
        # Ces avoirs doivent être sortis avec leur remboursement correspondant
        total_avoirs_affectes = sum(fac_info['avoir_affecte'] for fac_info in factures_avec_solde)
        if total_avoirs < total_avoirs_affectes + 0.01:
            # Tous les avoirs ont été affectés, rien à faire
            pass
        else:
            # Il reste des avoirs non affectés
            avoirs_non_affectes = total_avoirs - total_avoirs_affectes
            
            # Récupérer les remboursements de ce groupe
            remboursements_ce_groupe = remboursements_par_groupe.get(key, [])
            total_remboursements_groupe = sum(r['montant'] for r in remboursements_ce_groupe)
            
            # Pour chaque avoir non affecté, créer une ligne
            for avoir in avoirs_groupe:
                # Vérifier si cet avoir a été partiellement ou non affecté
                montant_avoir = avoir['montant']
                avoir_obj = avoir.get('avoir_obj')
                
                # Calculer la portion non affectée (simplifié: on traite proportionnellement)
                if total_avoirs > 0:
                    ratio_non_affecte = avoirs_non_affectes / total_avoirs
                    montant_avoir_non_affecte = montant_avoir * ratio_non_affecte
                else:
                    montant_avoir_non_affecte = 0
                
                if montant_avoir_non_affecte > 0.01:
                    # Calculer le remboursement correspondant
                    if total_avoirs > 0:
                        ratio_remb = montant_avoir / total_avoirs
                        montant_remboursement = total_remboursements_groupe * ratio_remb
                    else:
                        montant_remboursement = 0
                    
                    # Créer la ligne pour l'avoir non affecté
                    if avoir_obj is not None:
                        resultats.append({
                            'Date de facture': avoir_obj['Date'],
                            'N° de facture': avoir_obj['NumPiece'],
                            'N° compte fournisseur': compte_fournisseur,
                            'Nom du fournisseur': nom_fournisseur,
                            'Libellé de l\'opération': avoir_obj['Libelle'],
                            'Montant de la facture': 0,
                            'Avoir': montant_avoir_non_affecte,
                            'Montant facture net': -montant_avoir_non_affecte,
                            'Date de paiement': remboursements_ce_groupe[0]['date'] if remboursements_ce_groupe else None,
                            'Montant du paiement': -montant_remboursement if montant_remboursement > 0 else 0,
                            'OD': 0,
                            'Montant du paiement groupé': 0,
                            'Lettrage': lettrage_corrige,
                            'Lettrage corrigé': '',
                            'Solde': -montant_avoir_non_affecte + montant_remboursement
                        })

        groupes_traites.add(key)

    # Ajouter les reclassements OD lettrés
    # Ces OD ont MontantFacture > 0 et MontantMvt = 0 (ex: reclassement solde débiteur)
    # Deux cas :
    # 1. Si le groupe a des factures → ligne séparée avec OD
    # 2. Si le groupe N'A PAS de factures mais a des paiements → combiner OD + paiement
    for key, od_list in od_reclassements_par_groupe.items():
        compte_fournisseur, lettrage = key
        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")
        
        # Vérifier si ce groupe a des factures
        groupe_a_factures = key in factures_par_groupe and len(factures_par_groupe[key]) > 0
        
        # Récupérer les paiements de ce groupe
        paiements_groupe = paiements_par_groupe.get(key, [])
        
        for od_recl in od_list:
            montant_od_recl = od_recl['MontantFacture']
            
            if groupe_a_factures:
                # Cas 1 : groupe avec factures → ligne séparée
                resultats.append({
                    'Date de facture': od_recl['Date'],
                    'N° de facture': od_recl['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': od_recl['Libelle'],
                    'Montant de la facture': 0,
                    'Avoir': 0,
                    'Montant facture net': 0,
                    'Date de paiement': None,
                    'Montant du paiement': 0,
                    'OD': montant_od_recl,
                    'Montant du paiement groupé': 0,
                    'Lettrage': str(od_recl['Lettrage']).strip() if pd.notna(od_recl['Lettrage']) else '',
                    'Lettrage corrigé': '',
                    'Solde': montant_od_recl  # Reclassement augmente la dette
                })
            elif len(paiements_groupe) > 0:
                # Cas 2 : groupe SANS factures mais avec paiements → combiner OD + paiement
                # L'OD crée une dette qui est ensuite payée
                for paiement in paiements_groupe:
                    # Calculer la portion de l'OD correspondant à ce paiement (prorata si plusieurs)
                    total_paiements = sum(p['montant'] for p in paiements_groupe)
                    ratio = paiement['montant'] / total_paiements if total_paiements > 0 else 1
                    od_portion = montant_od_recl * ratio
                    
                    resultats.append({
                        'Date de facture': od_recl['Date'],
                        'N° de facture': od_recl['NumPiece'],
                        'N° compte fournisseur': compte_fournisseur,
                        'Nom du fournisseur': nom_fournisseur,
                        'Libellé de l\'opération': od_recl['Libelle'],
                        'Montant de la facture': 0,
                        'Avoir': 0,
                        'Montant facture net': 0,
                        'Date de paiement': paiement['date'],
                        'Montant du paiement': paiement['montant'],
                        'OD': od_portion,
                        'Montant du paiement groupé': total_paiements if len(paiements_groupe) > 1 else 0,
                        'Lettrage': str(od_recl['Lettrage']).strip() if pd.notna(od_recl['Lettrage']) else '',
                        'Lettrage corrigé': '',
                        'Solde': 0  # MF(0) - Paiement + OD = 0 car Paiement = OD
                    })
                # Marquer les paiements comme traités pour ne pas les réafficher
                paiements_par_groupe[key] = []
            else:
                # Cas 3 : groupe SANS factures ET SANS paiements → ligne séparée
                resultats.append({
                    'Date de facture': od_recl['Date'],
                    'N° de facture': od_recl['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': od_recl['Libelle'],
                    'Montant de la facture': 0,
                    'Avoir': 0,
                    'Montant facture net': 0,
                    'Date de paiement': None,
                    'Montant du paiement': 0,
                    'OD': montant_od_recl,
                    'Montant du paiement groupé': 0,
                    'Lettrage': str(od_recl['Lettrage']).strip() if pd.notna(od_recl['Lettrage']) else '',
                    'Lettrage corrigé': '',
                    'Solde': montant_od_recl
                })

    # Ajouter les avoirs lettrés dans des groupes SANS factures
    # Ces avoirs sont dans des groupes avec remboursements ou OD, mais sans factures
    # Cas typiques :
    # 1. Avoir + Remboursement (le fournisseur nous rembourse l'avoir)
    # 2. Avoir + OD (compensation entre avoir et OD)
    for key, avoirs_list in avoirs_par_groupe.items():
        compte_fournisseur, lettrage = key
        
        # Vérifier si ce groupe a des factures (déjà traité)
        if key in factures_par_groupe and len(factures_par_groupe[key]) > 0:
            continue  # Déjà traité avec les factures
        
        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")
        
        # Récupérer les remboursements de ce groupe
        remboursements_groupe = remboursements_par_groupe.get(key, [])
        
        # Calculer le total des avoirs du groupe
        total_avoirs_groupe = sum(avoir['montant'] for avoir in avoirs_list)
        
        # Calculer le total des remboursements du groupe
        total_remboursements = sum(remb['montant'] for remb in remboursements_groupe)
        
        for avoir in avoirs_list:
            avoir_obj = avoir['avoir_obj']
            montant_avoir = avoir['montant']
            
            if len(remboursements_groupe) > 0:
                # Cas 1 : Avoir + Remboursement
                # Le remboursement "paie" l'avoir (solde = 0)
                # Trouver le remboursement correspondant (prorata si plusieurs)
                ratio = montant_avoir / total_avoirs_groupe if total_avoirs_groupe > 0 else 1
                montant_remboursement = total_remboursements * ratio
                
                resultats.append({
                    'Date de facture': avoir_obj['Date'],
                    'N° de facture': avoir_obj['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': avoir_obj['Libelle'],
                    'Montant de la facture': 0,
                    'Avoir': montant_avoir,
                    'Montant facture net': -montant_avoir,
                    'Date de paiement': remboursements_groupe[0]['date'] if remboursements_groupe else None,
                    'Montant du paiement': -montant_remboursement,  # Négatif car remboursement
                    'OD': 0,
                    'Montant du paiement groupé': 0,
                    'Lettrage': lettrage,
                    'Lettrage corrigé': '',
                    'Solde': 0 if abs(montant_avoir - montant_remboursement) < 0.01 else -montant_avoir + montant_remboursement
                })
            else:
                # Cas 2 : Avoir sans remboursement dans le groupe
                # Vérifier s'il y a une OD qui compense
                od_groupe = dict_od_brut.get(key, 0)
                od_recl_groupe = od_reclassements_par_groupe.get(key, [])
                od_recl_total = sum(od['MontantFacture'] for od in od_recl_groupe)
                
                # NOTE: Les OD de od_reclassements_par_groupe sont déjà traitées séparément
                # On ne les ajoute PAS ici pour éviter le double comptage
                # On ajoute uniquement od_groupe (écarts de change MontantMvt > 0)
                if od_groupe > 0 or od_recl_total > 0:
                    # Avoir compensé par OD - l'OD reclassement est traitée séparément
                    # Le Solde ici ne doit PAS inclure l'OD (déjà comptée ailleurs)
                    resultats.append({
                        'Date de facture': avoir_obj['Date'],
                        'N° de facture': avoir_obj['NumPiece'],
                        'N° compte fournisseur': compte_fournisseur,
                        'Nom du fournisseur': nom_fournisseur,
                        'Libellé de l\'opération': avoir_obj['Libelle'],
                        'Montant de la facture': 0,
                        'Avoir': montant_avoir,
                        'Montant facture net': -montant_avoir,
                        'Date de paiement': None,
                        'Montant du paiement': 0,
                        'OD': 0,  # OD reclassement déjà comptée, on met 0 ici
                        'Montant du paiement groupé': 0,
                        'Lettrage': lettrage,
                        'Lettrage corrigé': '',
                        'Solde': -montant_avoir  # Solde = 0 - Avoir - 0 + 0 = -Avoir
                    })
                else:
                    # Avoir seul (non compensé) - ne devrait pas arriver normalement
                    resultats.append({
                        'Date de facture': avoir_obj['Date'],
                        'N° de facture': avoir_obj['NumPiece'],
                        'N° compte fournisseur': compte_fournisseur,
                        'Nom du fournisseur': nom_fournisseur,
                        'Libellé de l\'opération': avoir_obj['Libelle'],
                        'Montant de la facture': 0,
                        'Avoir': montant_avoir,
                        'Montant facture net': -montant_avoir,
                        'Date de paiement': None,
                        'Montant du paiement': 0,
                        'OD': 0,
                        'Montant du paiement groupé': 0,
                        'Lettrage': lettrage,
                        'Lettrage corrigé': '',
                        'Solde': -montant_avoir
                    })

    # Ajouter les factures non lettrées à la liste des factures avec solde
    for _, facture in factures.iterrows():
        compte = str(facture['Compte']).strip()
        lettrage = str(facture['LettrageCorrige']).strip()
        if not lettrage:
            nom_fournisseur = dict_fournisseurs.get(compte, "Fournisseur inconnu")
            factures_solde_restant.append({
                'facture': facture,
                'compte': compte,
                'nom_fournisseur': nom_fournisseur,
                'solde_restant': abs(facture['MontantFacture']),
                'lettrage_original': str(facture['Lettrage']).strip(),
                'lettrage_corrige_affiche': '',
                'od_signe': 0
            })

    # Trier toutes les factures avec solde par compte puis par date (les plus anciennes d'abord)
    factures_solde_restant = sorted(
        factures_solde_restant,
        key=lambda x: (x['compte'], x['facture']['Date'] if pd.notna(x['facture']['Date']) else pd.Timestamp.max)
    )

    # Identifier les paiements non lettrés (tous les mouvements != 0)
    paiements_non_lettres = paiements[
        (paiements['LettrageCorrige'] == '') &
        (paiements['MontantMvt'] != 0)
    ].copy()

    # Trier les paiements non lettrés par date
    paiements_non_lettres = paiements_non_lettres.sort_values('Date')

    # Suivre les affectations ET les montants restants des paiements
    affectations_paiements = {}  # clé = index dans factures_solde_restant
    paiements_restants = {}  # clé = (compte, numpiece, date, montant) → montant restant

    # Affecter les paiements non lettrés aux factures avec solde (même fournisseur, plus anciennes d'abord)
    for _, paiement in paiements_non_lettres.iterrows():
        compte_paiement = str(paiement['Compte']).strip()
        montant_paiement_initial = abs(paiement['MontantMvt'])
        montant_paiement_restant = montant_paiement_initial
        paiement_key = (compte_paiement, paiement['NumPiece'], paiement['Date'], paiement['MontantMvt'])

        for idx, fac_info in enumerate(factures_solde_restant):
            if montant_paiement_restant <= 0:
                break
            if fac_info['compte'] != compte_paiement:
                continue
            if fac_info['solde_restant'] <= 0:
                continue

            montant_a_affecter = min(montant_paiement_restant, fac_info['solde_restant'])

            if idx not in affectations_paiements:
                affectations_paiements[idx] = []
            affectations_paiements[idx].append({
                'paiement': paiement,
                'montant_affecte': montant_a_affecter
            })

            fac_info['solde_restant'] -= montant_a_affecter
            montant_paiement_restant -= montant_a_affecter

        # Stocker le montant restant UNIQUEMENT si le paiement a été PARTIELLEMENT affecté
        # (pas complètement non affecté)
        if montant_paiement_restant > 0.01 and montant_paiement_restant < montant_paiement_initial:
            paiements_restants[paiement_key] = {
                'paiement': paiement,
                'montant_restant': montant_paiement_restant
            }

    # Générer les lignes de résultat pour les factures avec solde
    for idx, fac_info in enumerate(factures_solde_restant):
        facture = fac_info['facture']
        compte_fournisseur = fac_info['compte']
        nom_fournisseur = fac_info['nom_fournisseur']
        lettrage_original = fac_info['lettrage_original']
        lettrage_corrige_affiche = fac_info['lettrage_corrige_affiche']
        od_signe = fac_info['od_signe']

        if idx in affectations_paiements:
            # La facture a reçu des paiements non lettrés
            for aff in affectations_paiements[idx]:
                paiement = aff['paiement']
                montant_paiement = aff['montant_affecte']

                resultats.append({
                    'Date de facture': facture['Date'],
                    'N° de facture': facture['NumPiece'],
                    'N° compte fournisseur': compte_fournisseur,
                    'Nom du fournisseur': nom_fournisseur,
                    'Libellé de l\'opération': facture['Libelle'],
                    'Montant de la facture': montant_paiement,
                    'Avoir': 0,
                    'Montant facture net': montant_paiement,
                    'Date de paiement': paiement['Date'],
                    'Montant du paiement': montant_paiement,
                    'OD': 0,
                    'Montant du paiement groupé': 0,
                    'Lettrage': lettrage_original,
                    'Lettrage corrigé': lettrage_corrige_affiche,
                    'Solde': 0
                })

        # Ajouter une ligne pour le solde restant si > 0
        if fac_info['solde_restant'] > 0.01:
            resultats.append({
                'Date de facture': facture['Date'],
                'N° de facture': facture['NumPiece'],
                'N° compte fournisseur': compte_fournisseur,
                'Nom du fournisseur': nom_fournisseur,
                'Libellé de l\'opération': facture['Libelle'],
                'Montant de la facture': fac_info['solde_restant'],
                'Avoir': 0,
                'Montant facture net': fac_info['solde_restant'],
                'Date de paiement': None,
                'Montant du paiement': 0,
                'OD': od_signe,
                'Montant du paiement groupé': 0,
                'Lettrage': lettrage_original,
                'Lettrage corrigé': lettrage_corrige_affiche,
                'Solde': fac_info['solde_restant'] + od_signe
            })

    # Ajouter les paiements sans lettrage qui n'ont pas été affectés OU partiellement affectés
    paiements_sans_lettrage = paiements[paiements['LettrageCorrige'] == '']

    paiements_deja_affectes = set()
    for idx, affectations in affectations_paiements.items():
        for aff in affectations:
            paiement = aff['paiement']
            paiements_deja_affectes.add((str(paiement['Compte']).strip(), paiement['NumPiece'], paiement['Date'], paiement['MontantMvt']))

    # 1. Ajouter les paiements qui n'ont PAS DU TOUT été affectés
    for _, paiement in paiements_sans_lettrage.iterrows():
        compte_fournisseur = str(paiement['Compte']).strip()
        paiement_key = (compte_fournisseur, paiement['NumPiece'], paiement['Date'], paiement['MontantMvt'])

        # Si le paiement a été affecté (totalement ou partiellement), on ne l'ajoute pas ici
        if paiement_key in paiements_deja_affectes:
            continue

        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")
        montant_mvt = paiement['MontantMvt']
        montant_paiement = abs(montant_mvt)

        # Solde = -MontantMvt (car paiement diminue la dette)
        # Si MontantMvt > 0 (avoir/remboursement), solde = -positif = négatif
        # Si MontantMvt < 0 (paiement), solde = -négatif = positif (erreur car on a payé sans facture)
        solde_paiement = -montant_mvt

        resultats.append({
            'Date de facture': None,
            'N° de facture': '',
            'N° compte fournisseur': compte_fournisseur,
            'Nom du fournisseur': nom_fournisseur,
            'Libellé de l\'opération': paiement['Libelle'],
            'Montant de la facture': 0,
            'Avoir': 0,
            'Montant facture net': 0,
            'Date de paiement': paiement['Date'],
            'Montant du paiement': montant_paiement,
            'OD': 0,
            'Montant du paiement groupé': 0,
            'Lettrage': '',
            'Lettrage corrigé': '',
            'Solde': solde_paiement
        })

    # 2. Ajouter les RELIQUATS des paiements partiellement affectés
    for paiement_key, paiement_info in paiements_restants.items():
        paiement = paiement_info['paiement']
        montant_restant = paiement_info['montant_restant']
        compte_fournisseur = str(paiement['Compte']).strip()
        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")

        # Solde = -montant_restant (crédit car on a trop payé)
        solde_reliquat = -montant_restant

        resultats.append({
            'Date de facture': None,
            'N° de facture': '',
            'N° compte fournisseur': compte_fournisseur,
            'Nom du fournisseur': nom_fournisseur,
            'Libellé de l\'opération': paiement['Libelle'] + ' (reliquat)',
            'Montant de la facture': 0,
            'Avoir': 0,
            'Montant facture net': 0,
            'Date de paiement': paiement['Date'],
            'Montant du paiement': montant_restant,
            'OD': 0,
            'Montant du paiement groupé': 0,
            'Lettrage': '',
            'Lettrage corrigé': '',
            'Solde': solde_reliquat
        })

    # Ajouter les OD non lettrés (journaux autres que achat/banque, sans lettrage)
    od_non_lettres = grand_livre_df[
        (grand_livre_df['Compte'].astype(str).str.startswith('4411')) &
        (~grand_livre_df['Journal'].isin(journaux_connus)) &
        ((grand_livre_df['LettrageCorrige'].str.strip() == '') | (grand_livre_df['LettrageCorrige'].isna())) &
        ((grand_livre_df['MontantFacture'] != 0) | (grand_livre_df['MontantMvt'] != 0))
    ].copy()

    for _, od in od_non_lettres.iterrows():
        compte_fournisseur = str(od['Compte']).strip()
        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")
        
        # OD = MontantFacture - MontantMvt (ColG - ColF)
        montant_od = od['MontantFacture'] - od['MontantMvt']

        resultats.append({
            'Date de facture': od['Date'],
            'N° de facture': od['NumPiece'],
            'N° compte fournisseur': compte_fournisseur,
            'Nom du fournisseur': nom_fournisseur,
            'Libellé de l\'opération': od['Libelle'],
            'Montant de la facture': 0,
            'Avoir': 0,
            'Montant facture net': 0,
            'Date de paiement': None,
            'Montant du paiement': 0,
            'OD': montant_od,
            'Montant du paiement groupé': 0,
            'Lettrage': '',
            'Lettrage corrigé': '',
            'Solde': montant_od
        })

    # Ajouter les avoirs non lettrés (sans lettrage)
    for avoir in avoirs_non_lettres:
        compte_fournisseur = str(avoir['Compte']).strip()
        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")
        montant_avoir = avoir['MontantMvt']
        
        resultats.append({
            'Date de facture': None,
            'N° de facture': avoir['NumPiece'],
            'N° compte fournisseur': compte_fournisseur,
            'Nom du fournisseur': nom_fournisseur,
            'Libellé de l\'opération': avoir['Libelle'],
            'Montant de la facture': 0,
            'Avoir': montant_avoir,
            'Montant facture net': -montant_avoir,
            'Date de paiement': avoir['Date'],
            'Montant du paiement': 0,
            'OD': 0,
            'Montant du paiement groupé': 0,
            'Lettrage': '',
            'Lettrage corrigé': '',
            'Solde': -montant_avoir
        })

    # Ajouter les remboursements fournisseurs non lettrés
    for remb in remboursements_non_lettres.iterrows():
        remboursement = remb[1]
        compte_fournisseur = str(remboursement['Compte']).strip()
        nom_fournisseur = dict_fournisseurs.get(compte_fournisseur, "Fournisseur inconnu")
        montant_remboursement = remboursement['MontantFacture']
        
        resultats.append({
            'Date de facture': None,
            'N° de facture': remboursement['NumPiece'],
            'N° compte fournisseur': compte_fournisseur,
            'Nom du fournisseur': nom_fournisseur,
            'Libellé de l\'opération': remboursement['Libelle'],
            'Montant de la facture': 0,
            'Avoir': 0,
            'Montant facture net': 0,
            'Date de paiement': remboursement['Date'],
            'Montant du paiement': -montant_remboursement,  # Négatif car nous remboursons
            'OD': 0,
            'Montant du paiement groupé': 0,
            'Lettrage': '',
            'Lettrage corrigé': '',
            'Solde': montant_remboursement  # Positif car dette envers le fournisseur
        })

    df_resultats = pd.DataFrame(resultats)

    for col in ['Date de facture', 'Date de paiement']:
        if col in df_resultats.columns:
            df_resultats[col] = pd.to_datetime(df_resultats[col], errors='coerce')

    if len(df_resultats) > 0:
        df_resultats = df_resultats.sort_values(
            by=['N° compte fournisseur', 'Date de facture'],
            ascending=[True, True],
            na_position='last'
        ).reset_index(drop=True)

        df_resultats['N° de facture'] = df_resultats['N° de facture'].astype(str)
        df_resultats['N° compte fournisseur'] = df_resultats['N° compte fournisseur'].astype(str)
        df_resultats['Nom du fournisseur'] = df_resultats['Nom du fournisseur'].astype(str)
        df_resultats['Libellé de l\'opération'] = df_resultats['Libellé de l\'opération'].astype(str)
        df_resultats['Lettrage'] = df_resultats['Lettrage'].astype(str)
        df_resultats['Lettrage corrigé'] = df_resultats['Lettrage corrigé'].astype(str)

    return df_resultats

def export_to_excel(df):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Calculer la ligne de total
        colonnes_montants = ['Montant de la facture', 'Avoir', 'Montant facture net',
                           'Montant du paiement', 'OD', 'Montant du paiement groupé', 'Solde']

        totaux = {'Date de facture': 'TOTAL'}
        for col in df.columns:
            if col in colonnes_montants:
                totaux[col] = df[col].sum()
            elif col == 'Date de facture':
                totaux[col] = 'TOTAL'
            else:
                totaux[col] = ''

        # Créer un DataFrame avec la ligne total en premier
        df_total = pd.DataFrame([totaux])
        df_export = pd.concat([df_total, df], ignore_index=True)

        df_export.to_excel(writer, sheet_name='Rapprochement', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Rapprochement']

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        total_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')

        cell_alignment = Alignment(horizontal='center', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Appliquer le style aux en-têtes (ligne 1)
        for col_idx in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Appliquer le style à la ligne TOTAL (ligne 2)
        for col_idx in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border

        # Colonnes de montants (indices basés sur l'ordre des colonnes)
        col_indices_montants = []
        for i, col_name in enumerate(df_export.columns, 1):
            if col_name in colonnes_montants:
                col_indices_montants.append(i)

        # Colonnes de dates
        col_indices_dates = []
        for i, col_name in enumerate(df_export.columns, 1):
            if col_name in ['Date de facture', 'Date de paiement']:
                col_indices_dates.append(i)

        # Formater toutes les cellules
        for row in range(2, len(df_export) + 2):
            for col_idx in range(1, len(df_export.columns) + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.border = thin_border

                if col_idx in col_indices_dates:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = cell_alignment
                elif col_idx in col_indices_montants:
                    # Format avec séparateur de milliers et "-" pour les zéros
                    # Ce format affiche "-" pour 0 tout en gardant la valeur numérique
                    cell.number_format = '#,##0.00;-#,##0.00;"-"'
                    cell.alignment = number_alignment
                else:
                    cell.alignment = cell_alignment

        # Ajuster la largeur des colonnes
        column_widths = {
            'Date de facture': 12,
            'N° de facture': 15,
            'N° compte fournisseur': 12,
            'Nom du fournisseur': 25,
            'Libellé de l\'opération': 30,
            'Montant de la facture': 15,
            'Avoir': 12,
            'Montant facture net': 15,
            'Date de paiement': 12,
            'Montant du paiement': 15,
            'OD': 12,
            'Montant du paiement groupé': 15,
            'Lettrage': 10,
            'Lettrage corrigé': 12,
            'Solde': 12
        }

        for i, col_name in enumerate(df_export.columns, 1):
            col_letter = get_column_letter(i)
            width = column_widths.get(col_name, 15)
            worksheet.column_dimensions[col_letter].width = width

        # Hauteur de la ligne d'en-tête
        worksheet.row_dimensions[1].height = 40

        # Figer les volets en cellule E3 (lignes 1-2 et colonnes A-D figées)
        worksheet.freeze_panes = 'E3'

    output.seek(0)
    return output

# Interface principale
if grand_livre_file and balance_file:
    try:
        with st.spinner("Chargement des fichiers..."):
            # Lire les bytes des fichiers pour le cache
            gl_bytes = grand_livre_file.read()
            grand_livre_file.seek(0)
            balance_bytes = balance_file.read()
            balance_file.seek(0)

            grand_livre_df = load_grand_livre(gl_bytes)
            balance_df, has_header = load_balance(balance_bytes)

        dict_fournisseurs, col_compte, col_nom = creer_dict_fournisseurs(balance_df, has_header)

        st.success("Fichiers chargés avec succès !")

        # Métriques dans des cartes
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown('<div class="metric-container">', unsafe_allow_html=True)
            st.metric("Lignes Grand Livre", f"{len(grand_livre_df):,}".replace(',', ' '))
            st.markdown('</div>', unsafe_allow_html=True)
        with col2:
            st.markdown('<div class="metric-container">', unsafe_allow_html=True)
            st.metric("Fournisseurs", f"{len(dict_fournisseurs):,}".replace(',', ' '))
            st.markdown('</div>', unsafe_allow_html=True)
        with col3:
            st.markdown('<div class="metric-container">', unsafe_allow_html=True)
            st.metric("Journaux Achat", len(journaux_achat))
            st.markdown('</div>', unsafe_allow_html=True)

        with st.expander("Aperçu du Grand Livre", expanded=False):
            st.dataframe(grand_livre_df.head(20), use_container_width=True)

        with st.expander("Aperçu de la Balance", expanded=False):
            st.info(f"Colonnes detectees - Compte: **{col_compte}**, Nom: **{col_nom}**")
            st.dataframe(balance_df.head(20), use_container_width=True)

            if dict_fournisseurs:
                st.markdown("**Mapping Compte -> Fournisseur (10 premiers)**")
                mapping_preview = list(dict_fournisseurs.items())[:10]
                mapping_df = pd.DataFrame(mapping_preview, columns=['N° Compte', 'Nom Fournisseur'])
                st.dataframe(mapping_df, use_container_width=True)

        comptes_gl = set(grand_livre_df[grand_livre_df['Compte'].str.startswith('4411')]['Compte'].unique())
        comptes_trouves = comptes_gl.intersection(set(dict_fournisseurs.keys()))
        comptes_non_trouves = comptes_gl - comptes_trouves

        if comptes_non_trouves:
            with st.expander(f"{len(comptes_non_trouves)} compte(s) non trouve(s) dans la balance", expanded=False):
                st.warning("Ces comptes du Grand Livre n'ont pas de correspondance :")
                st.write(sorted(list(comptes_non_trouves))[:20])

        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        # Bouton centré
        col_btn_left, col_btn_center, col_btn_right = st.columns([1, 2, 1])
        with col_btn_center:
            if st.button("Lancer le rapprochement", type="primary", use_container_width=True):
                with st.spinner("Traitement en cours..."):
                    resultats_df = traiter_rapprochement(
                        grand_livre_df,
                        dict_fournisseurs,
                        journaux_achat,
                        journaux_banque
                    )

                st.success(f"Rapprochement termine ! **{len(resultats_df)} lignes** generees")

                # Calcul du solde attendu depuis le grand livre
                gl_4411 = grand_livre_df[grand_livre_df['Compte'].astype(str).str.startswith('4411')]
                total_g = gl_4411['MontantFacture'].sum()  # Total colonne G
                total_f = gl_4411['MontantMvt'].sum()       # Total colonne F
                solde_attendu = total_g - total_f

                # Solde calculé depuis les résultats
                solde_calcule = resultats_df['Solde'].sum()

                # Statistiques
                col1, col2, col3, col4 = st.columns(4)

                factures_payees = len(resultats_df[resultats_df['Date de paiement'].notna()])
                factures_non_payees = len(resultats_df[resultats_df['Date de paiement'].isna()]) - len(resultats_df[resultats_df['Date de facture'].isna()])
                paiements_sans_facture = len(resultats_df[resultats_df['Date de facture'].isna()])

                with col1:
                    st.metric("Factures payees", factures_payees)
                with col2:
                    st.metric("Factures non payees", factures_non_payees)
                with col3:
                    st.metric("Paiements sans facture", paiements_sans_facture)
                with col4:
                    total_montant = resultats_df['Montant de la facture'].sum()
                    st.metric("Total factures", f"{total_montant:,.2f} MAD".replace(',', ' ').replace('.', ','))

                # Vérification du solde
                st.markdown("### Verification du solde")
                col_v1, col_v2, col_v3 = st.columns(3)
                with col_v1:
                    st.metric("Total Colonne G (Factures)", f"{total_g:,.2f}".replace(',', ' ').replace('.', ','))
                with col_v2:
                    st.metric("Total Colonne F (Mouvements)", f"{total_f:,.2f}".replace(',', ' ').replace('.', ','))
                with col_v3:
                    st.metric("Solde attendu (G - F)", f"{solde_attendu:,.2f}".replace(',', ' ').replace('.', ','))

                ecart = abs(solde_calcule - solde_attendu)
                if ecart < 0.01:
                    st.success(f"✅ Solde calculé: **{solde_calcule:,.2f}** MAD - Conforme au grand livre !".replace(',', ' ').replace('.', ','))
                else:
                    st.warning(f"⚠️ Solde calculé: **{solde_calcule:,.2f}** MAD - Écart de **{ecart:,.2f}** MAD".replace(',', ' ').replace('.', ','))

                st.markdown("### Resultats du rapprochement")

                # Affichage formaté
                display_df = resultats_df.copy()
                display_df['Date de facture'] = display_df['Date de facture'].dt.strftime('%d/%m/%Y')
                display_df['Date de paiement'] = display_df['Date de paiement'].dt.strftime('%d/%m/%Y')

                colonnes_montants = ['Montant de la facture', 'Avoir', 'Montant facture net',
                                    'Montant du paiement', 'OD', 'Montant du paiement groupé', 'Solde']
                for col in colonnes_montants:
                    if col in display_df.columns:
                        display_df[col] = display_df[col].apply(
                            lambda x: "-" if x == 0 else f"{x:,.2f}".replace(',', ' ').replace('.', ',') if pd.notna(x) else ''
                        )

                display_df = display_df.fillna('')

                st.dataframe(
                    display_df,
                    use_container_width=True,
                    height=400
                )

                # Téléchargement
                excel_file = export_to_excel(resultats_df)

                col_dl_left, col_dl_center, col_dl_right = st.columns([1, 2, 1])
                with col_dl_center:
                    st.download_button(
                        label="Telecharger le fichier Excel",
                        data=excel_file,
                        file_name=f"rapprochement_delais_paiement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )

    except Exception as e:
        st.error(f"Erreur lors du traitement : {str(e)}")
        st.exception(e)

else:
    col_info_left, col_info_center, col_info_right = st.columns([1, 2, 1])
    with col_info_center:
        st.info("Veuillez charger le Grand Livre et la Balance des Comptes pour commencer.")

    with st.expander("Instructions d'utilisation", expanded=True):
        st.markdown("""
        ### Comment utiliser cet outil ?

        1. **Chargez le Grand Livre** (fichier Excel sans en-tete)
           - Colonne A : Date | Colonne B : Journal | Colonne C : N° Compte
           - Colonne D : N° Piece | Colonne E : Libelle | Colonne F : Montant Mvt
           - Colonne G : Montant Facture | Colonne I : Lettrage

        2. **Chargez la Balance des Comptes**
           - Doit contenir le N° de compte et le nom du fournisseur

        3. **Configurez les journaux** d'achat et de banque

        4. **Cliquez sur "Lancer le rapprochement"**

        5. **Telechargez le fichier Excel** avec les resultats

        ---

        ### Regles de rapprochement

        | Type | Criteres |
        |------|----------|
        | **Factures** | Comptes 4411* + Journal d'achat + Montant facture != 0 |
        | **Paiements** | Journal de banque + Montant mouvement != 0 |
        | **Rapprochement** | Meme lettrage + Meme compte fournisseur |
        | **OD** | Journaux autres que achat/banque avec lettrage |
        """)

# Footer
st.markdown("""
<div class="footer">
    <strong>Expert-Comptable Morocco</strong> | Outil de declaration des delais de paiement
</div>
""", unsafe_allow_html=True)
